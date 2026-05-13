"""Sky Strike Remote Config build script.

Reads XLSX workbooks in remote_config/ and emits:
  - assets/config/defaults.json           (committed; consumed by Flutter)
  - tools/out/firebase_remote_config_template.json (gitignored; upload payload)

Usage:
    python tools/build_config.py            # build defaults.json + template
    python tools/build_config.py --init     # (re)create the 6 seed workbooks
    python tools/build_config.py --init --force   # overwrite existing workbooks
    python tools/build_config.py --validate # validate only, no output

Layout (relative to project root /Users/idanc83/Projects/Sky Strike):
    remote_config/*.xlsx          (sibling of skystrike/)
    skystrike/tools/build_config.py
    skystrike/assets/config/defaults.json
    skystrike/tools/out/firebase_remote_config_template.json
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


# --------------------------------------------------------------------------
# Paths
# --------------------------------------------------------------------------

SCRIPT_PATH = Path(__file__).resolve()
SKYSTRIKE_ROOT = SCRIPT_PATH.parent.parent              # .../Sky Strike/skystrike
OUTER_ROOT = SKYSTRIKE_ROOT.parent                      # .../Sky Strike
REMOTE_CONFIG_DIR = OUTER_ROOT / "remote_config"
DEFAULTS_JSON = SKYSTRIKE_ROOT / "assets" / "config" / "defaults.json"
TEMPLATE_JSON = SKYSTRIKE_ROOT / "tools" / "out" / "firebase_remote_config_template.json"


# --------------------------------------------------------------------------
# Styles
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


# --------------------------------------------------------------------------
# Namespace schema versions (Python <-> Dart must agree)
# --------------------------------------------------------------------------

SCHEMA_VERSIONS: dict[str, int] = {
    "difficulty": 1,
    "drops": 1,
    "economy": 1,
    "progression": 1,
    "flags": 1,
    "experiments": 1,
}


# --------------------------------------------------------------------------
# Errors
# --------------------------------------------------------------------------

class BuildError(Exception):
    """Raised when validation fails. Carries file + sheet + row context."""


def fail(workbook: str, sheet: str | None, row: int | None, message: str) -> None:
    """Raise a BuildError with file/sheet/row prefix."""
    loc = f"[{workbook}"
    if sheet:
        loc += f":{sheet}"
    if row is not None:
        loc += f"!row{row}"
    loc += "]"
    raise BuildError(f"{loc} {message}")


# --------------------------------------------------------------------------
# Workbook helpers
# --------------------------------------------------------------------------

def _style_header(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 12)
    ws.freeze_panes = "A2"


def _add_bool_validation(ws: Worksheet, col_letter: str, last_row: int) -> None:
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    dv.add(f"{col_letter}2:{col_letter}{max(last_row, 2)}")
    ws.add_data_validation(dv)


def _add_prob_validation(ws: Worksheet, col_letter: str, last_row: int) -> None:
    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv.add(f"{col_letter}2:{col_letter}{max(last_row, 2)}")
    ws.add_data_validation(dv)


def _write_meta(wb: Workbook, namespace: str, description: str) -> None:
    ws = wb.create_sheet("_meta")
    _style_header(ws, ["key", "value"])
    rows = [
        ("schema_version", SCHEMA_VERSIONS[namespace]),
        ("description", description),
        ("last_edited_by", ""),
    ]
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)


def _percent_format(ws: Worksheet, col_letter: str, last_row: int) -> None:
    for r in range(2, last_row + 1):
        ws[f"{col_letter}{r}"].number_format = "0.00%"


def _two_decimal_format(ws: Worksheet, col_letter: str, last_row: int) -> None:
    for r in range(2, last_row + 1):
        ws[f"{col_letter}{r}"].number_format = "0.00"


def _currency_format(ws: Worksheet, col_letter: str, last_row: int) -> None:
    for r in range(2, last_row + 1):
        ws[f"{col_letter}{r}"].number_format = "#,##0"


# --------------------------------------------------------------------------
# --init : generate seed workbooks
# --------------------------------------------------------------------------

def _init_difficulty(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_meta(wb, "difficulty",
                "Per-(world, wave) difficulty curves and per-world enemy base stats")

    ws = wb.create_sheet("wave_curves")
    headers = ["world", "wave", "hp_mult", "speed_mult", "spawn_count",
               "elites_allowed", "enemy_fire", "is_boss", "_notes"]
    _style_header(ws, headers)
    seed_waves = [
        (1, 1, 1.00, 1.00,  8, "FALSE", "FALSE", "FALSE", ""),
        (1, 2, 1.05, 1.00,  9, "FALSE", "FALSE", "FALSE", ""),
        (1, 3, 1.10, 1.00, 10, "FALSE", "FALSE", "FALSE", ""),
        (1, 4, 1.15, 1.02, 11, "FALSE", "TRUE",  "FALSE", ""),
        (1, 5, 1.20, 1.04, 12, "FALSE", "TRUE",  "FALSE", ""),
        (1, 6, 1.25, 1.06, 13, "TRUE",  "TRUE",  "FALSE", ""),
        (1, 7, 1.30, 1.08, 14, "TRUE",  "TRUE",  "FALSE", ""),
        (1, 8, 1.35, 1.10, 15, "TRUE",  "TRUE",  "FALSE", ""),
        (1, 9, 1.40, 1.10, 15, "TRUE",  "TRUE",  "FALSE", ""),
        (1,10, 1.60, 1.15, 18, "TRUE",  "TRUE",  "TRUE",  "boss wave"),
    ]
    for r, row in enumerate(seed_waves, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    last = 1 + len(seed_waves)
    _two_decimal_format(ws, "C", last)
    _two_decimal_format(ws, "D", last)
    _add_bool_validation(ws, "F", last)
    _add_bool_validation(ws, "G", last)
    _add_bool_validation(ws, "H", last)

    ws2 = wb.create_sheet("enemy_scaling")
    _style_header(ws2, ["world", "base_hp", "base_speed", "fire_rate", "_notes"])
    ws2.append([1, 10, 80, 0.0, ""])

    wb.save(path)


def _init_drops(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_meta(wb, "drops", "Drop rate buckets and power-up distribution")

    ws = wb.create_sheet("rates")
    _style_header(ws, ["world", "wave_min", "wave_max", "hp_prob",
                       "powerup_prob", "gem_prob", "_notes"])
    rows = [
        (1, 1, 5, 0.15, 0.10, 0.02, "HP boosted for tutorial waves"),
        (1, 6, 10, 0.08, 0.12, 0.03, ""),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    last = 1 + len(rows)
    for col in ("D", "E", "F"):
        _percent_format(ws, col, last)
        _add_prob_validation(ws, col, last)

    ws2 = wb.create_sheet("powerup_distribution")
    _style_header(ws2, ["powerup_id", "weight", "category", "_notes"])
    pdrops = [
        ("bomb",          0.25, "collectible", ""),
        ("laser",         0.15, "collectible", ""),
        ("magnet",        0.20, "collectible", ""),
        ("ghost",         0.15, "collectible", ""),
        ("freeze",        0.10, "collectible", ""),
        ("rapid_fire",    0.00, "instant",     ""),
        ("shield",        0.00, "instant",     "Shield NEVER drops"),
        ("split_shot",    0.10, "instant",     ""),
        ("speed_boost",   0.05, "instant",     ""),
        ("drone_wingman", 0.00, "instant",     ""),
    ]
    for r, row in enumerate(pdrops, start=2):
        for c, val in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=val)
    _two_decimal_format(ws2, "B", 1 + len(pdrops))

    ws3 = wb.create_sheet("streak_boost")
    _style_header(ws3, ["key", "value", "_notes"])
    streak = [
        ("trigger_streak_modulo", 5,   ""),
        ("hp_mult",               2.0, ""),
        ("powerup_mult",          1.5, ""),
        ("gem_mult",              1.5, ""),
    ]
    for r, row in enumerate(streak, start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(row=r, column=c, value=val)

    wb.save(path)


def _init_economy(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_meta(wb, "economy",
                "Revive pricing brackets, shop catalog, daily deals pool")

    ws = wb.create_sheet("revive_brackets")
    _style_header(ws, ["wave_min", "wave_max", "cost_gems", "_notes"])
    for row in [(1, 3, 5, ""), (4, 6, 10, ""), (7, 9, 15, "")]:
        ws.append(list(row))

    ws2 = wb.create_sheet("revive_global")
    _style_header(ws2, ["key", "value", "_notes"])
    for row in [("boss_cost_gems", 20, ""), ("max_revives_per_run", 1, "")]:
        ws2.append(list(row))

    ws3 = wb.create_sheet("shop_jets")
    _style_header(ws3, ["jet_id", "currency", "amount", "sku", "_notes"])
    for row in [
        ("scout",   "gems", 120, "", ""),
        ("phantom", "gems", 300, "", ""),
    ]:
        ws3.append(list(row))

    ws4 = wb.create_sheet("shop_bundles")
    _style_header(ws4, ["bundle_id", "currency", "amount", "sku", "_notes"])
    for row in [
        ("starter_pack",   "usd",  299, "com.skystrike.starter", ""),
        ("powerup_bundle", "gems",  80, "", ""),
    ]:
        ws4.append(list(row))

    ws5 = wb.create_sheet("daily_deals_pool")
    _style_header(ws5, ["pool_id", "_notes"])
    for row in [("scout", ""), ("phantom", ""), ("powerup_bundle", "")]:
        ws5.append(list(row))

    wb.save(path)


def _init_progression(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_meta(wb, "progression", "XP curve and per-level rewards")

    ws = wb.create_sheet("xp_curve")
    _style_header(ws, ["level", "xp_cumulative", "_notes"])
    seed = [
        (1, 0, ""),
        (2, 100, ""),
        (3, 250, ""),
        (4, 450, ""),
        (5, 700, ""),
        (6, 1000, ""),
        (7, 1400, ""),
        (8, 1900, ""),
        (9, 2500, ""),
        (10, 3200, ""),
    ]
    for r, row in enumerate(seed, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    _currency_format(ws, "B", 1 + len(seed))

    ws2 = wb.create_sheet("xp_curve_config")
    _style_header(ws2, ["key", "value", "_notes"])
    ws2.append(["level_cap", 100, ""])

    ws3 = wb.create_sheet("level_rewards")
    _style_header(ws3, ["level", "coins", "gems", "powerups", "jet", "_notes"])
    rewards = [
        (2, 100, 5, "", "", ""),
        (5, 500, 20, "bomb;laser", "", ""),
        (10, 1000, 50, "", "scout", ""),
    ]
    for r, row in enumerate(rewards, start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(row=r, column=c, value=val)
    _currency_format(ws3, "B", 1 + len(rewards))
    _currency_format(ws3, "C", 1 + len(rewards))

    wb.save(path)


def _init_flags(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_meta(wb, "flags", "Feature flags and kill switches")

    ws = wb.create_sheet("feature_flags")
    _style_header(ws, ["flag_key", "value", "_notes"])
    flags = [
        ("world_2_enabled",             "FALSE", ""),
        ("world_3_enabled",             "FALSE", ""),
        ("operation_challenge_enabled", "TRUE",  ""),
        ("iap_remove_ads_enabled",      "TRUE",  ""),
        ("rewarded_ads_enabled",        "TRUE",  ""),
        ("leaderboard_enabled",         "FALSE", ""),
    ]
    for r, row in enumerate(flags, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    _add_bool_validation(ws, "B", 1 + len(flags))

    ws2 = wb.create_sheet("kill_switches")
    _style_header(ws2, ["switch_key", "value", "_notes"])
    switches = [
        ("kill_iap",         "FALSE", ""),
        ("kill_ads",         "FALSE", ""),
        ("kill_leaderboard", "TRUE",  ""),
    ]
    for r, row in enumerate(switches, start=2):
        for c, val in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=val)
    _add_bool_validation(ws2, "B", 1 + len(switches))

    wb.save(path)


def _init_experiments(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_meta(wb, "experiments",
                "A/B experiment default assignments. In production, "
                "Firebase conditions override these per-user.")

    ws = wb.create_sheet("ab_assignment")
    _style_header(ws, ["experiment_key", "default_variant",
                       "available_variants", "_notes"])
    rows = [
        ("first_purchase_offer", "control", "control;variant_a;variant_b", ""),
        ("hud_layout",           "control", "control;compact;expanded",    ""),
        ("tutorial_length",      "short",   "short;long",                  ""),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)


INIT_BUILDERS: dict[str, Callable[[Path], None]] = {
    "difficulty.xlsx":  _init_difficulty,
    "drops.xlsx":       _init_drops,
    "economy.xlsx":     _init_economy,
    "progression.xlsx": _init_progression,
    "flags.xlsx":       _init_flags,
    "experiments.xlsx": _init_experiments,
}


def cmd_init(force: bool) -> int:
    REMOTE_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    created = 0
    skipped = 0
    for name, builder in INIT_BUILDERS.items():
        path = REMOTE_CONFIG_DIR / name
        if path.exists() and not force:
            print(f"skip   {name} (exists; pass --force to overwrite)")
            skipped += 1
            continue
        builder(path)
        print(f"wrote  {name}")
        created += 1
    print(f"\n--init complete. wrote={created} skipped={skipped} "
          f"dir={REMOTE_CONFIG_DIR}")
    return 0


# --------------------------------------------------------------------------
# Reading + validation
# --------------------------------------------------------------------------

@dataclass
class SheetRows:
    headers: list[str]
    rows: list[dict[str, Any]]


def _read_sheet(wb_path: Path, wb: Workbook, sheet_name: str) -> SheetRows:
    if sheet_name not in wb.sheetnames:
        fail(wb_path.name, sheet_name, None, "sheet missing")
    ws = wb[sheet_name]
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not headers_row:
        fail(wb_path.name, sheet_name, 1, "header row is empty")
    headers = [str(h) if h is not None else "" for h in headers_row]
    out: list[dict[str, Any]] = []
    for r_idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if raw is None or all(v is None or v == "" for v in raw):
            continue
        row: dict[str, Any] = {}
        for h, v in zip(headers, raw):
            if not h or h == "_notes":
                continue
            row[h] = v
        row["__row__"] = r_idx
        out.append(row)
    return SheetRows(headers=headers, rows=out)


def _require_meta(wb_path: Path, wb: Workbook, namespace: str) -> dict[str, Any]:
    meta_rows = _read_sheet(wb_path, wb, "_meta").rows
    kv: dict[str, Any] = {r["key"]: r.get("value") for r in meta_rows if "key" in r}
    expected = SCHEMA_VERSIONS[namespace]
    seen = kv.get("schema_version")
    if seen is None:
        fail(wb_path.name, "_meta", None, "schema_version missing")
    try:
        seen_int = int(seen)
    except (TypeError, ValueError):
        fail(wb_path.name, "_meta", None,
             f"schema_version must be int, got {seen!r}")
    if seen_int != expected:
        fail(wb_path.name, "_meta", None,
             f"schema_version={seen_int} but build script expects {expected}. "
             f"Bump SCHEMA_VERSIONS in build_config.py and the matching Dart "
             f"schema's supportedSchemaVersion together.")
    return kv


def _as_bool(v: Any, wb_path: Path, sheet: str, row: int, col: str) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        if v.strip().upper() == "TRUE":
            return True
        if v.strip().upper() == "FALSE":
            return False
    fail(wb_path.name, sheet, row,
         f"column {col!r} must be TRUE/FALSE, got {v!r}")


def _as_int(v: Any, wb_path: Path, sheet: str, row: int, col: str) -> int:
    if isinstance(v, bool):
        fail(wb_path.name, sheet, row, f"column {col!r} expected int, got bool")
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str) and v.strip().lstrip("-").isdigit():
        return int(v.strip())
    fail(wb_path.name, sheet, row, f"column {col!r} expected int, got {v!r}")


def _as_float(v: Any, wb_path: Path, sheet: str, row: int, col: str) -> float:
    if isinstance(v, bool):
        fail(wb_path.name, sheet, row, f"column {col!r} expected number, got bool")
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            pass
    fail(wb_path.name, sheet, row, f"column {col!r} expected number, got {v!r}")


def _as_str(v: Any, wb_path: Path, sheet: str, row: int, col: str,
            allow_blank: bool = False) -> str:
    if v is None:
        if allow_blank:
            return ""
        fail(wb_path.name, sheet, row, f"column {col!r} is blank")
    s = str(v).strip()
    if not s and not allow_blank:
        fail(wb_path.name, sheet, row, f"column {col!r} is blank")
    return s


# --------------------------------------------------------------------------
# Namespace builders (XLSX -> payload dict)
# --------------------------------------------------------------------------

def _build_difficulty(wb_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(wb_path, data_only=False)
    _require_meta(wb_path, wb, "difficulty")

    waves_sheet = _read_sheet(wb_path, wb, "wave_curves")
    waves: dict[str, dict[str, Any]] = {}
    for r in waves_sheet.rows:
        rn = r["__row__"]
        world = _as_int(r.get("world"), wb_path, "wave_curves", rn, "world")
        wave = _as_int(r.get("wave"), wb_path, "wave_curves", rn, "wave")
        key = f"{world}_{wave}"
        if key in waves:
            fail(wb_path.name, "wave_curves", rn,
                 f"duplicate (world,wave) {key}")
        waves[key] = {
            "hp_mult":        _as_float(r.get("hp_mult"), wb_path, "wave_curves", rn, "hp_mult"),
            "speed_mult":     _as_float(r.get("speed_mult"), wb_path, "wave_curves", rn, "speed_mult"),
            "spawn_count":    _as_int(r.get("spawn_count"), wb_path, "wave_curves", rn, "spawn_count"),
            "elites_allowed": _as_bool(r.get("elites_allowed"), wb_path, "wave_curves", rn, "elites_allowed"),
            "enemy_fire":     _as_bool(r.get("enemy_fire"), wb_path, "wave_curves", rn, "enemy_fire"),
            "is_boss":        _as_bool(r.get("is_boss"), wb_path, "wave_curves", rn, "is_boss"),
        }

    scaling_sheet = _read_sheet(wb_path, wb, "enemy_scaling")
    worlds: dict[str, dict[str, Any]] = {}
    for r in scaling_sheet.rows:
        rn = r["__row__"]
        world = _as_int(r.get("world"), wb_path, "enemy_scaling", rn, "world")
        wkey = str(world)
        if wkey in worlds:
            fail(wb_path.name, "enemy_scaling", rn, f"duplicate world {wkey}")
        worlds[wkey] = {
            "base_hp":    _as_int(r.get("base_hp"), wb_path, "enemy_scaling", rn, "base_hp"),
            "base_speed": _as_int(r.get("base_speed"), wb_path, "enemy_scaling", rn, "base_speed"),
            "fire_rate":  _as_float(r.get("fire_rate"), wb_path, "enemy_scaling", rn, "fire_rate"),
        }

    return {
        "wave_curves":   {"schema_version": SCHEMA_VERSIONS["difficulty"], "waves": waves},
        "enemy_scaling": {"schema_version": SCHEMA_VERSIONS["difficulty"], "worlds": worlds},
    }


def _build_drops(wb_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(wb_path, data_only=False)
    _require_meta(wb_path, wb, "drops")

    rates_sheet = _read_sheet(wb_path, wb, "rates")
    by_world: dict[int, list[tuple[int, int, dict[str, Any], int]]] = {}
    buckets: list[dict[str, Any]] = []
    for r in rates_sheet.rows:
        rn = r["__row__"]
        world = _as_int(r.get("world"), wb_path, "rates", rn, "world")
        wmin = _as_int(r.get("wave_min"), wb_path, "rates", rn, "wave_min")
        wmax = _as_int(r.get("wave_max"), wb_path, "rates", rn, "wave_max")
        if wmin > wmax:
            fail(wb_path.name, "rates", rn,
                 f"wave_min {wmin} > wave_max {wmax}")
        bucket = {
            "world": world,
            "wave_min": wmin,
            "wave_max": wmax,
            "hp_prob":      _as_float(r.get("hp_prob"), wb_path, "rates", rn, "hp_prob"),
            "powerup_prob": _as_float(r.get("powerup_prob"), wb_path, "rates", rn, "powerup_prob"),
            "gem_prob":     _as_float(r.get("gem_prob"), wb_path, "rates", rn, "gem_prob"),
        }
        buckets.append(bucket)
        by_world.setdefault(world, []).append((wmin, wmax, bucket, rn))

    for world, items in by_world.items():
        items.sort(key=lambda t: t[0])
        cover: list[tuple[int, int]] = [(wmin, wmax) for wmin, wmax, _, _ in items]
        expected = 1
        for wmin, wmax in cover:
            if wmin != expected:
                fail(wb_path.name, "rates", None,
                     f"world {world}: gap or overlap; expected wave_min={expected} "
                     f"got {wmin}")
            expected = wmax + 1
        if expected != 11:
            fail(wb_path.name, "rates", None,
                 f"world {world}: buckets must cover waves 1..10 "
                 f"(last wave_max+1={expected}, want 11)")

    dist_sheet = _read_sheet(wb_path, wb, "powerup_distribution")
    distribution: dict[str, float] = {}
    weight_sum = 0.0
    for r in dist_sheet.rows:
        rn = r["__row__"]
        pid = _as_str(r.get("powerup_id"), wb_path, "powerup_distribution", rn, "powerup_id")
        weight = _as_float(r.get("weight"), wb_path, "powerup_distribution", rn, "weight")
        if weight < 0:
            fail(wb_path.name, "powerup_distribution", rn,
                 f"weight must be >= 0, got {weight}")
        if pid in distribution:
            fail(wb_path.name, "powerup_distribution", rn,
                 f"duplicate powerup_id {pid!r}")
        distribution[pid] = weight
        weight_sum += weight
    if weight_sum <= 0:
        fail(wb_path.name, "powerup_distribution", None,
             "weights must sum to > 0")

    streak_sheet = _read_sheet(wb_path, wb, "streak_boost")
    streak_kv: dict[str, Any] = {}
    for r in streak_sheet.rows:
        rn = r["__row__"]
        k = _as_str(r.get("key"), wb_path, "streak_boost", rn, "key")
        streak_kv[k] = r.get("value")
    required = ("trigger_streak_modulo", "hp_mult", "powerup_mult", "gem_mult")
    for k in required:
        if k not in streak_kv:
            fail(wb_path.name, "streak_boost", None, f"missing key {k!r}")
    streak_boost = {
        "schema_version": SCHEMA_VERSIONS["drops"],
        "trigger_streak_modulo": int(streak_kv["trigger_streak_modulo"]),
        "hp_mult":      float(streak_kv["hp_mult"]),
        "powerup_mult": float(streak_kv["powerup_mult"]),
        "gem_mult":     float(streak_kv["gem_mult"]),
    }

    return {
        "rates": {
            "schema_version": SCHEMA_VERSIONS["drops"],
            "buckets": buckets,
        },
        "powerup_distribution": {
            "schema_version": SCHEMA_VERSIONS["drops"],
            "distribution": distribution,
        },
        "streak_boost": streak_boost,
    }


def _build_economy(wb_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(wb_path, data_only=False)
    _require_meta(wb_path, wb, "economy")

    revive_sheet = _read_sheet(wb_path, wb, "revive_brackets")
    brackets: list[dict[str, Any]] = []
    for r in revive_sheet.rows:
        rn = r["__row__"]
        wmin = _as_int(r.get("wave_min"), wb_path, "revive_brackets", rn, "wave_min")
        wmax = _as_int(r.get("wave_max"), wb_path, "revive_brackets", rn, "wave_max")
        cost = _as_int(r.get("cost_gems"), wb_path, "revive_brackets", rn, "cost_gems")
        if wmin > wmax:
            fail(wb_path.name, "revive_brackets", rn,
                 f"wave_min {wmin} > wave_max {wmax}")
        brackets.append({"wave_min": wmin, "wave_max": wmax, "cost_gems": cost})
    brackets.sort(key=lambda b: b["wave_min"])
    expected = 1
    for b in brackets:
        if b["wave_min"] != expected:
            fail(wb_path.name, "revive_brackets", None,
                 f"gap or overlap; expected wave_min={expected} got {b['wave_min']}")
        expected = b["wave_max"] + 1
    if expected != 10:
        fail(wb_path.name, "revive_brackets", None,
             f"brackets must cover waves 1..9 (last wave_max+1={expected}, want 10)")

    global_sheet = _read_sheet(wb_path, wb, "revive_global")
    gkv: dict[str, Any] = {}
    for r in global_sheet.rows:
        rn = r["__row__"]
        k = _as_str(r.get("key"), wb_path, "revive_global", rn, "key")
        gkv[k] = r.get("value")
    for k in ("boss_cost_gems", "max_revives_per_run"):
        if k not in gkv:
            fail(wb_path.name, "revive_global", None, f"missing key {k!r}")

    def _parse_shop(sheet_name: str, id_col: str) -> dict[str, dict[str, Any]]:
        sheet = _read_sheet(wb_path, wb, sheet_name)
        out: dict[str, dict[str, Any]] = {}
        for r in sheet.rows:
            rn = r["__row__"]
            item_id = _as_str(r.get(id_col), wb_path, sheet_name, rn, id_col)
            currency = _as_str(r.get("currency"), wb_path, sheet_name, rn, "currency")
            if currency not in ("gems", "coins", "usd"):
                fail(wb_path.name, sheet_name, rn,
                     f"currency must be gems/coins/usd, got {currency!r}")
            amount = _as_int(r.get("amount"), wb_path, sheet_name, rn, "amount")
            sku = _as_str(r.get("sku"), wb_path, sheet_name, rn, "sku", allow_blank=True)
            if item_id in out:
                fail(wb_path.name, sheet_name, rn, f"duplicate id {item_id!r}")
            out[item_id] = {"currency": currency, "amount": amount, "sku": sku}
        return out

    jets = _parse_shop("shop_jets", "jet_id")
    bundles = _parse_shop("shop_bundles", "bundle_id")

    pool_sheet = _read_sheet(wb_path, wb, "daily_deals_pool")
    pool: list[str] = []
    for r in pool_sheet.rows:
        rn = r["__row__"]
        pid = _as_str(r.get("pool_id"), wb_path, "daily_deals_pool", rn, "pool_id")
        pool.append(pid)

    revive_pricing = {
        "schema_version": SCHEMA_VERSIONS["economy"],
        "wave_brackets": brackets,
        "boss_cost_gems":      int(gkv["boss_cost_gems"]),
        "max_revives_per_run": int(gkv["max_revives_per_run"]),
    }
    shop_prices = {
        "schema_version": SCHEMA_VERSIONS["economy"],
        "jets": jets,
        "bundles": bundles,
        "daily_deals_pool": pool,
    }
    return {"revive_pricing": revive_pricing, "shop_prices": shop_prices}


def _build_progression(wb_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(wb_path, data_only=False)
    _require_meta(wb_path, wb, "progression")

    curve_sheet = _read_sheet(wb_path, wb, "xp_curve")
    levels: list[tuple[int, int]] = []
    for r in curve_sheet.rows:
        rn = r["__row__"]
        level = _as_int(r.get("level"), wb_path, "xp_curve", rn, "level")
        xp = _as_int(r.get("xp_cumulative"), wb_path, "xp_curve", rn, "xp_cumulative")
        levels.append((level, xp))
    levels.sort(key=lambda t: t[0])
    if not levels or levels[0][0] != 1:
        fail(wb_path.name, "xp_curve", None, "level column must start at 1")
    for i, (lvl, xp) in enumerate(levels):
        if lvl != i + 1:
            fail(wb_path.name, "xp_curve", None,
                 f"levels must be 1..N with no gaps; got {lvl} at index {i}")
        if i > 0 and xp <= levels[i - 1][1]:
            fail(wb_path.name, "xp_curve", None,
                 f"xp_cumulative must be strictly increasing at level {lvl}")
    if levels[0][1] != 0:
        fail(wb_path.name, "xp_curve", None,
             "level 1 must have xp_cumulative=0")
    xp_cumulative = [xp for _, xp in levels]

    cfg_sheet = _read_sheet(wb_path, wb, "xp_curve_config")
    cfg_kv: dict[str, Any] = {}
    for r in cfg_sheet.rows:
        rn = r["__row__"]
        k = _as_str(r.get("key"), wb_path, "xp_curve_config", rn, "key")
        cfg_kv[k] = r.get("value")
    if "level_cap" not in cfg_kv:
        fail(wb_path.name, "xp_curve_config", None, "missing key 'level_cap'")
    level_cap = int(cfg_kv["level_cap"])

    rewards_sheet = _read_sheet(wb_path, wb, "level_rewards")
    rewards: dict[str, dict[str, Any]] = {}
    for r in rewards_sheet.rows:
        rn = r["__row__"]
        level = _as_int(r.get("level"), wb_path, "level_rewards", rn, "level")
        coins = _as_int(r.get("coins"), wb_path, "level_rewards", rn, "coins")
        gems = _as_int(r.get("gems"), wb_path, "level_rewards", rn, "gems")
        pups_raw = _as_str(r.get("powerups"), wb_path, "level_rewards", rn,
                           "powerups", allow_blank=True)
        jet = _as_str(r.get("jet"), wb_path, "level_rewards", rn, "jet",
                      allow_blank=True)
        pups = [p.strip() for p in pups_raw.split(";") if p.strip()]
        key = str(level)
        if key in rewards:
            fail(wb_path.name, "level_rewards", rn,
                 f"duplicate level {level}")
        rewards[key] = {
            "coins": coins,
            "gems": gems,
            "powerups": pups,
            "jet": jet,
        }

    return {
        "xp_curve": {
            "schema_version": SCHEMA_VERSIONS["progression"],
            "level_cap": level_cap,
            "xp_cumulative": xp_cumulative,
        },
        "level_rewards": {
            "schema_version": SCHEMA_VERSIONS["progression"],
            "rewards": rewards,
        },
    }


def _build_flags(wb_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(wb_path, data_only=False)
    _require_meta(wb_path, wb, "flags")

    def _parse_flags(sheet_name: str, key_col: str) -> dict[str, bool]:
        sheet = _read_sheet(wb_path, wb, sheet_name)
        out: dict[str, bool] = {}
        for r in sheet.rows:
            rn = r["__row__"]
            k = _as_str(r.get(key_col), wb_path, sheet_name, rn, key_col)
            v = _as_bool(r.get("value"), wb_path, sheet_name, rn, "value")
            if k in out:
                fail(wb_path.name, sheet_name, rn, f"duplicate key {k!r}")
            out[k] = v
        return out

    return {
        "feature_flags": {
            "schema_version": SCHEMA_VERSIONS["flags"],
            "flags": _parse_flags("feature_flags", "flag_key"),
        },
        "kill_switches": {
            "schema_version": SCHEMA_VERSIONS["flags"],
            "flags": _parse_flags("kill_switches", "switch_key"),
        },
    }


def _build_experiments(wb_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(wb_path, data_only=False)
    _require_meta(wb_path, wb, "experiments")

    sheet = _read_sheet(wb_path, wb, "ab_assignment")
    experiments: dict[str, str] = {}
    for r in sheet.rows:
        rn = r["__row__"]
        ek = _as_str(r.get("experiment_key"), wb_path, "ab_assignment", rn, "experiment_key")
        dv = _as_str(r.get("default_variant"), wb_path, "ab_assignment", rn, "default_variant")
        av = _as_str(r.get("available_variants"), wb_path, "ab_assignment", rn, "available_variants")
        legal = [v.strip() for v in av.split(";") if v.strip()]
        if dv not in legal:
            fail(wb_path.name, "ab_assignment", rn,
                 f"default_variant {dv!r} not in available_variants {legal}")
        if ek in experiments:
            fail(wb_path.name, "ab_assignment", rn, f"duplicate experiment_key {ek!r}")
        experiments[ek] = dv

    return {
        "ab_assignment": {
            "schema_version": SCHEMA_VERSIONS["experiments"],
            "experiments": experiments,
        },
    }


# --------------------------------------------------------------------------
# Key mapping (namespace+name -> Firebase key)
# --------------------------------------------------------------------------

# Maps (namespace, payload_key_in_builder_output) -> Firebase Remote Config key
KEY_MAP: list[tuple[str, str, str, str]] = [
    # (namespace, payload_name, firebase_key, description)
    ("difficulty",  "wave_curves",          "difficulty__wave_curves__v1",
        "Per-(world, wave) difficulty curves"),
    ("difficulty",  "enemy_scaling",        "difficulty__enemy_scaling__v1",
        "Per-world enemy base stats"),
    ("drops",       "rates",                "drops__rates__v1",
        "Drop rate buckets per (world, wave range)"),
    ("drops",       "powerup_distribution", "drops__powerup_distribution__v1",
        "Weighted power-up distribution"),
    ("drops",       "streak_boost",         "drops__streak_boost__v1",
        "Failure-streak drop boost"),
    ("economy",     "revive_pricing",       "economy__revive_pricing__v1",
        "Revive pricing brackets"),
    ("economy",     "shop_prices",          "economy__shop_prices__v1",
        "Shop catalog (jets, bundles, daily deals)"),
    ("progression", "xp_curve",             "progression__xp_curve__v1",
        "XP cumulative curve + level cap"),
    ("progression", "level_rewards",        "progression__level_rewards__v1",
        "Per-level reward grants"),
    ("flags",       "feature_flags",        "flags__feature_flags__v1",
        "Feature flags"),
    ("flags",       "kill_switches",        "flags__kill_switches__v1",
        "Kill switches (TRUE = killed)"),
    ("experiments", "ab_assignment",        "experiments__ab_assignment__v1",
        "A/B default variant assignments"),
]


NAMESPACE_BUILDERS: dict[str, Callable[[Path], dict[str, dict[str, Any]]]] = {
    "difficulty":  _build_difficulty,
    "drops":       _build_drops,
    "economy":     _build_economy,
    "progression": _build_progression,
    "flags":       _build_flags,
    "experiments": _build_experiments,
}


def _build_all() -> dict[str, dict[str, Any]]:
    """Returns: {namespace: {payload_name: payload_dict, ...}}"""
    out: dict[str, dict[str, Any]] = {}
    for namespace, builder in NAMESPACE_BUILDERS.items():
        path = REMOTE_CONFIG_DIR / f"{namespace}.xlsx"
        if not path.exists():
            fail(path.name, None, None,
                 f"workbook not found at {path}. Run --init first.")
        out[namespace] = builder(path)
    return out


def _flatten_payloads(all_payloads: dict[str, dict[str, Any]]) -> dict[str, str]:
    """Map firebase_key -> JSON-stringified payload."""
    flat: dict[str, str] = {}
    for namespace, payload_name, fkey, _desc in KEY_MAP:
        if namespace not in all_payloads:
            raise BuildError(f"missing namespace in build output: {namespace}")
        if payload_name not in all_payloads[namespace]:
            raise BuildError(
                f"missing payload {payload_name!r} in namespace {namespace}")
        payload = all_payloads[namespace][payload_name]
        flat[fkey] = json.dumps(payload, separators=(",", ":"), sort_keys=True)
    return flat


def cmd_build(validate_only: bool) -> int:
    all_payloads = _build_all()
    flat = _flatten_payloads(all_payloads)

    if validate_only:
        print(f"OK  {len(flat)} keys validated, "
              f"namespaces={sorted(SCHEMA_VERSIONS.keys())}")
        return 0

    DEFAULTS_JSON.parent.mkdir(parents=True, exist_ok=True)
    TEMPLATE_JSON.parent.mkdir(parents=True, exist_ok=True)

    defaults_obj = dict(sorted(flat.items()))
    DEFAULTS_JSON.write_text(
        json.dumps(defaults_obj, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    template = {
        "parameters": {
            fkey: {
                "defaultValue": {"value": flat[fkey]},
                "valueType": "STRING",
                "description": desc,
            }
            for _, _, fkey, desc in KEY_MAP
        },
        "version": {
            "description": "Generated from remote_config/ workbooks by build_config.py",
        },
    }
    TEMPLATE_JSON.write_text(
        json.dumps(template, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    print(f"Wrote {len(flat)} keys, {len(SCHEMA_VERSIONS)} namespaces, "
          f"schema_versions={SCHEMA_VERSIONS}")
    print(f"  -> {DEFAULTS_JSON.relative_to(SKYSTRIKE_ROOT)}")
    print(f"  -> {TEMPLATE_JSON.relative_to(SKYSTRIKE_ROOT)}")
    return 0


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def _parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="build_config.py",
        description="Build Sky Strike Remote Config defaults + Firebase template.",
    )
    p.add_argument("--init", action="store_true",
                   help="(re)create seed workbooks in remote_config/")
    p.add_argument("--force", action="store_true",
                   help="with --init: overwrite existing workbooks")
    p.add_argument("--validate", action="store_true",
                   help="read and validate workbooks; do not write output")
    return p.parse_args(argv)


def main(argv: list[str]) -> int:
    args = _parse_args(argv)
    try:
        if args.init:
            return cmd_init(force=args.force)
        return cmd_build(validate_only=args.validate)
    except BuildError as e:
        print(f"ERROR  {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
